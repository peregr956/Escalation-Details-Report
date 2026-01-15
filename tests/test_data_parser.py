"""Unit tests for data_parser.py

This module provides comprehensive test coverage for the Excel data parsing
functions that load and parse incident data from Excel files.
"""
import pytest
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import List

from data_parser import (
    Incident,
    COLUMN_MAPPING,
    DATETIME_FIELDS,
    DURATION_FIELDS,
    parse_datetime,
    parse_duration,
    load_excel_file,
    load_multiple_periods,
    get_period_date_range,
    validate_excel_structure,
)


# =============================================================================
# Fixtures
# =============================================================================

@pytest.fixture
def sample_incident():
    """Return a fully populated Incident for testing."""
    return Incident(
        incident_id=12345,
        incident_url="https://example.com/incident/12345",
        vendor_incident_id="VND-001",
        vendor_incident_url="https://vendor.com/VND-001",
        incident_title="Test Security Alert",
        organization="Test Client Corp",
        product="Test EDR Product",
        deployment_status="Active",
        initial_escalation_method="Playbook",
        playbook_url="https://example.com/playbook/test",
        current_status="Closed",
        cs_soc_verdict="True Positive",
        current_priority="3 - HIGH",
        current_category="Malware",
        created_datetime_utc=datetime(2025, 6, 15, 10, 0, 0),
        created_datetime_local=datetime(2025, 6, 15, 5, 0, 0),
        escalated_datetime_utc=datetime(2025, 6, 15, 10, 30, 0),
        escalated_datetime_local=datetime(2025, 6, 15, 5, 30, 0),
        closed_datetime_utc=datetime(2025, 6, 15, 12, 0, 0),
        closed_datetime_local=datetime(2025, 6, 15, 7, 0, 0),
        escalation_paths="Primary Path",
        notification_groups="Security Team",
        assigned_users="analyst@criticalstart.com",
        touched_by="analyst@criticalstart.com, client@company.com",
        closed_by="analyst@criticalstart.com",
        cs_soc_last_comment="Threat contained",
        customer_last_comment="Verified by IT",
        response_action="Block IP",
        action_target="192.168.1.100",
        target_type="IP Address",
        action_provider="Firewall",
        executed_date=datetime(2025, 6, 15, 11, 0, 0),
        executed_by="system",
        response_action_status="Success",
        cs_soc_ttr=timedelta(hours=1, minutes=30),
        cs_soc_ttd=timedelta(minutes=30),
        customer_ttr=timedelta(hours=2),
        customer_ttd=timedelta(minutes=45),
        mitre_tactic_id="TA0003",
        mitre_tactic_name="Persistence",
        mitre_technique_id="T1547",
        mitre_technique_name="Boot or Logon Autostart Execution",
        vendor_severity="High",
    )


# =============================================================================
# Tests for Incident dataclass
# =============================================================================

class TestIncidentDataclass:
    """Tests for the Incident dataclass."""
    
    def test_default_values(self):
        """Test that all fields default to None."""
        incident = Incident()
        
        assert incident.incident_id is None
        assert incident.organization is None
        assert incident.product is None
        assert incident.current_status is None
        assert incident.cs_soc_ttr is None
    
    def test_field_assignment(self, sample_incident):
        """Test that fields can be assigned correctly."""
        assert sample_incident.incident_id == 12345
        assert sample_incident.organization == "Test Client Corp"
        assert sample_incident.product == "Test EDR Product"
        assert sample_incident.current_priority == "3 - HIGH"
        assert sample_incident.cs_soc_verdict == "True Positive"
    
    def test_datetime_fields(self, sample_incident):
        """Test that datetime fields hold datetime objects."""
        assert isinstance(sample_incident.created_datetime_utc, datetime)
        assert isinstance(sample_incident.escalated_datetime_utc, datetime)
        assert isinstance(sample_incident.closed_datetime_utc, datetime)
    
    def test_timedelta_fields(self, sample_incident):
        """Test that duration fields hold timedelta objects."""
        assert isinstance(sample_incident.cs_soc_ttr, timedelta)
        assert isinstance(sample_incident.cs_soc_ttd, timedelta)
        
        # Verify duration values
        assert sample_incident.cs_soc_ttr.total_seconds() == 90 * 60  # 1h 30m
        assert sample_incident.cs_soc_ttd.total_seconds() == 30 * 60  # 30m


# =============================================================================
# Tests for COLUMN_MAPPING
# =============================================================================

class TestColumnMapping:
    """Tests for the COLUMN_MAPPING dictionary."""
    
    def test_required_columns_exist(self):
        """Test that required columns are mapped."""
        required_columns = [
            "Incident Id",
            "Organization",
            "Product",
            "Current Status",
            "Current Priority",
            "CS SOC Verdict",
        ]
        
        for col in required_columns:
            assert col in COLUMN_MAPPING, f"Missing required column: {col}"
    
    def test_datetime_columns_mapped(self):
        """Test that datetime columns are properly mapped."""
        datetime_columns = [
            "Created Datetime UTC",
            "Escalated Datetime UTC",
            "Closed Datetime UTC",
        ]
        
        for col in datetime_columns:
            assert col in COLUMN_MAPPING
            assert COLUMN_MAPPING[col] in DATETIME_FIELDS
    
    def test_duration_columns_mapped(self):
        """Test that duration columns are properly mapped."""
        duration_columns = [
            "CS SOC TTR (hh:mm)",
            "CS SOC TTD (hh:mm)",
        ]
        
        for col in duration_columns:
            assert col in COLUMN_MAPPING
            assert COLUMN_MAPPING[col] in DURATION_FIELDS
    
    def test_alternate_formats_exist(self):
        """Test that alternate column name formats are supported."""
        # Uppercase ID variations
        assert "Incident ID" in COLUMN_MAPPING
        assert "Vendor Incident ID" in COLUMN_MAPPING
        
        # Alternate timestamp formats
        assert "Created Datetime (UTC)" in COLUMN_MAPPING
        assert "Escalated Datetime (UTC)" in COLUMN_MAPPING


# =============================================================================
# Tests for parse_datetime
# =============================================================================

class TestParseDatetime:
    """Tests for the parse_datetime function."""
    
    def test_none_input(self):
        """Test None returns None."""
        assert parse_datetime(None) is None
    
    def test_datetime_passthrough(self):
        """Test datetime objects are returned as-is."""
        dt = datetime(2025, 6, 15, 10, 30)
        result = parse_datetime(dt)
        assert result == dt
    
    def test_iso_format_string(self):
        """Test ISO format string parsing."""
        result = parse_datetime("2025-06-15T10:30:00")
        assert result == datetime(2025, 6, 15, 10, 30)
    
    def test_iso_format_with_microseconds(self):
        """Test ISO format with microseconds."""
        result = parse_datetime("2025-06-15T10:30:00.123456")
        assert result is not None
        assert result.year == 2025
        assert result.month == 6
        assert result.day == 15
    
    def test_iso_format_with_timezone(self):
        """Test ISO format with timezone (stripped)."""
        result = parse_datetime("2025-06-15T10:30:00+00:00")
        assert result is not None
        assert result.year == 2025
    
    def test_iso_format_with_z(self):
        """Test ISO format with Z suffix."""
        result = parse_datetime("2025-06-15T10:30:00Z")
        assert result is not None
        assert result.year == 2025
    
    def test_standard_format(self):
        """Test standard datetime format."""
        result = parse_datetime("2025-06-15 10:30:00")
        assert result == datetime(2025, 6, 15, 10, 30)
    
    def test_us_format(self):
        """Test US format (MM/DD/YYYY)."""
        result = parse_datetime("06/15/2025 10:30:00")
        assert result == datetime(2025, 6, 15, 10, 30)
    
    def test_invalid_format(self):
        """Test invalid format returns None."""
        result = parse_datetime("not a date")
        assert result is None
    
    def test_empty_string(self):
        """Test empty string returns None."""
        result = parse_datetime("")
        # Empty string should return None (not parse successfully)
        assert result is None


# =============================================================================
# Tests for parse_duration
# =============================================================================

class TestParseDuration:
    """Tests for the parse_duration function."""
    
    def test_none_input(self):
        """Test None returns None."""
        assert parse_duration(None) is None
    
    def test_timedelta_passthrough(self):
        """Test timedelta objects are returned as-is."""
        td = timedelta(hours=1, minutes=30)
        result = parse_duration(td)
        assert result == td
    
    def test_hhmm_format(self):
        """Test hh:mm format parsing."""
        result = parse_duration("1:30")
        assert result == timedelta(hours=1, minutes=30)
    
    def test_two_digit_hours(self):
        """Test two-digit hours format."""
        result = parse_duration("12:45")
        assert result == timedelta(hours=12, minutes=45)
    
    def test_single_digit_hour(self):
        """Test single-digit hour format."""
        result = parse_duration("2:15")
        assert result == timedelta(hours=2, minutes=15)
    
    def test_zero_hours(self):
        """Test zero hours format."""
        result = parse_duration("0:45")
        assert result == timedelta(minutes=45)
    
    def test_with_whitespace(self):
        """Test with surrounding whitespace."""
        result = parse_duration("  1:30  ")
        assert result == timedelta(hours=1, minutes=30)
    
    def test_invalid_format(self):
        """Test invalid format returns None."""
        result = parse_duration("not a duration")
        assert result is None
    
    def test_invalid_hhmm(self):
        """Test malformed hh:mm returns None."""
        result = parse_duration("1:3")  # Minutes should be 2 digits
        assert result is None


# =============================================================================
# Tests for get_period_date_range
# =============================================================================

class TestGetPeriodDateRange:
    """Tests for the get_period_date_range function."""
    
    def test_empty_list(self):
        """Test empty list returns (None, None)."""
        start, end = get_period_date_range([])
        assert start is None
        assert end is None
    
    def test_single_incident(self):
        """Test with single incident."""
        incidents = [
            Incident(
                created_datetime_utc=datetime(2025, 6, 15, 10, 0),
                escalated_datetime_utc=datetime(2025, 6, 15, 10, 30),
            )
        ]
        
        start, end = get_period_date_range(incidents)
        
        assert start == datetime(2025, 6, 15, 10, 0)
        assert end == datetime(2025, 6, 15, 10, 30)
    
    def test_multiple_incidents(self):
        """Test with multiple incidents spanning a range."""
        incidents = [
            Incident(created_datetime_utc=datetime(2025, 6, 1, 10, 0)),
            Incident(created_datetime_utc=datetime(2025, 6, 15, 10, 0)),
            Incident(created_datetime_utc=datetime(2025, 6, 30, 10, 0)),
        ]
        
        start, end = get_period_date_range(incidents)
        
        assert start == datetime(2025, 6, 1, 10, 0)
        assert end == datetime(2025, 6, 30, 10, 0)
    
    def test_no_dates(self):
        """Test incidents without dates returns (None, None)."""
        incidents = [
            Incident(organization="Test"),
            Incident(organization="Test2"),
        ]
        
        start, end = get_period_date_range(incidents)
        assert start is None
        assert end is None


# =============================================================================
# Integration test with temporary Excel file
# =============================================================================

class TestLoadExcelFile:
    """Tests for the load_excel_file function."""
    
    @pytest.fixture
    def temp_excel_file(self, tmp_path):
        """Create a temporary Excel file for testing."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Write headers
        headers = [
            "Incident Id", "Organization", "Product", 
            "Current Status", "Current Priority", "CS SOC Verdict",
            "Created Datetime UTC", "CS SOC TTR (hh:mm)"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write sample data rows
        sample_data = [
            [1, "Client A", "EDR Product", "Closed", "3 - HIGH", "True Positive", 
             datetime(2025, 6, 15, 10, 0), "1:30"],
            [2, "Client A", "Firewall", "Closed", "5 - MEDIUM", "False Positive",
             datetime(2025, 6, 15, 14, 0), "2:00"],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to temp file
        file_path = tmp_path / "test_data.xlsx"
        wb.save(file_path)
        
        return file_path
    
    def test_load_excel_file(self, temp_excel_file):
        """Test loading an Excel file."""
        incidents = load_excel_file(temp_excel_file)
        
        assert len(incidents) == 2
        assert incidents[0].incident_id == 1
        assert incidents[0].organization == "Client A"
        assert incidents[0].product == "EDR Product"
        assert incidents[0].current_status == "Closed"
        assert incidents[0].current_priority == "3 - HIGH"
    
    def test_load_nonexistent_file(self):
        """Test loading a nonexistent file raises FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            load_excel_file(Path("/nonexistent/file.xlsx"))
    
    def test_datetime_parsing(self, temp_excel_file):
        """Test that datetime fields are parsed correctly."""
        incidents = load_excel_file(temp_excel_file)
        
        assert incidents[0].created_datetime_utc is not None
        assert isinstance(incidents[0].created_datetime_utc, datetime)
    
    def test_duration_parsing(self, temp_excel_file):
        """Test that duration fields are parsed correctly."""
        incidents = load_excel_file(temp_excel_file)
        
        assert incidents[0].cs_soc_ttr is not None
        assert isinstance(incidents[0].cs_soc_ttr, timedelta)
        assert incidents[0].cs_soc_ttr == timedelta(hours=1, minutes=30)


class TestLoadMultiplePeriods:
    """Tests for the load_multiple_periods function."""
    
    @pytest.fixture
    def temp_excel_files(self, tmp_path):
        """Create multiple temporary Excel files for testing."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        files = []
        
        for period in range(3):
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Write headers
            headers = ["Incident Id", "Organization", "Current Status", "Current Priority"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            ws.cell(row=2, column=1, value=period * 10 + 1)
            ws.cell(row=2, column=2, value="Test Client")
            ws.cell(row=2, column=3, value="Closed")
            ws.cell(row=2, column=4, value="High")
            
            file_path = tmp_path / f"period_{period}.xlsx"
            wb.save(file_path)
            files.append(file_path)
        
        return files
    
    def test_load_multiple_periods(self, temp_excel_files):
        """Test loading multiple period files."""
        all_periods, client_name = load_multiple_periods(temp_excel_files)
        
        assert len(all_periods) == 3
        assert client_name == "Test Client"
    
    def test_empty_file_list(self):
        """Test empty file list raises ValueError."""
        with pytest.raises(ValueError, match="At least one Excel file"):
            load_multiple_periods([])


class TestValidateExcelStructure:
    """Tests for the validate_excel_structure function."""
    
    @pytest.fixture
    def valid_excel_file(self, tmp_path):
        """Create a valid Excel file for testing."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Write all expected headers
        headers = list(COLUMN_MAPPING.keys())[:10]  # First 10 columns
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write a data row
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="https://example.com")
        
        file_path = tmp_path / "valid.xlsx"
        wb.save(file_path)
        
        return file_path
    
    def test_valid_file(self, valid_excel_file):
        """Test validation of a valid file."""
        messages = validate_excel_structure(valid_excel_file)
        
        # Should find some data rows
        assert any("Found" in msg and "data rows" in msg for msg in messages)
    
    def test_nonexistent_file(self, tmp_path):
        """Test validation of nonexistent file."""
        messages = validate_excel_structure(tmp_path / "nonexistent.xlsx")
        
        assert any("not found" in msg.lower() for msg in messages)
    
    def test_empty_file(self, tmp_path):
        """Test validation of file with no data rows."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Only headers, no data
        headers = ["Incident Id", "Organization", "Current Status", "Current Priority"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)
        
        messages = validate_excel_structure(file_path)
        
        assert any("No data rows found" in msg for msg in messages)


# =============================================================================
# Tests for DATETIME_FIELDS and DURATION_FIELDS constants
# =============================================================================

class TestFieldConstants:
    """Tests for field constant sets."""
    
    def test_datetime_fields_complete(self):
        """Test that all expected datetime fields are defined."""
        expected = {
            "created_datetime_utc",
            "created_datetime_local",
            "escalated_datetime_utc",
            "escalated_datetime_local",
            "closed_datetime_utc",
            "closed_datetime_local",
            "executed_date",
        }
        
        for field in expected:
            assert field in DATETIME_FIELDS, f"Missing datetime field: {field}"
    
    def test_duration_fields_complete(self):
        """Test that all expected duration fields are defined."""
        expected = {
            "cs_soc_ttr",
            "cs_soc_ttd",
            "customer_ttr",
            "customer_ttd",
        }
        
        for field in expected:
            assert field in DURATION_FIELDS, f"Missing duration field: {field}"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
