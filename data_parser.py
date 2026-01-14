"""Data parser module for loading and parsing Excel incident data.

This module handles loading Excel files conforming to the 45-column incident
data structure and parsing them into typed Incident dataclass records.
"""
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional, Tuple
import re


@dataclass
class Incident:
    """Represents a single incident record from the Excel export."""
    
    # Identifiers
    incident_id: Optional[int] = None
    incident_url: Optional[str] = None
    vendor_incident_id: Optional[str] = None
    vendor_incident_url: Optional[str] = None
    incident_title: Optional[str] = None
    
    # Organization & Product
    organization: Optional[str] = None
    product: Optional[str] = None
    deployment_status: Optional[str] = None
    
    # Escalation Info
    initial_escalation_method: Optional[str] = None
    playbook_url: Optional[str] = None
    
    # Status & Classification
    current_status: Optional[str] = None
    cs_soc_verdict: Optional[str] = None
    current_priority: Optional[str] = None
    current_category: Optional[str] = None
    
    # Timestamps (stored as datetime objects)
    created_datetime_utc: Optional[datetime] = None
    created_datetime_local: Optional[datetime] = None
    last_updated_datetime_utc: Optional[datetime] = None
    last_updated_datetime_local: Optional[datetime] = None
    escalated_datetime_utc: Optional[datetime] = None
    escalated_datetime_local: Optional[datetime] = None
    closed_datetime_utc: Optional[datetime] = None
    closed_datetime_local: Optional[datetime] = None
    
    # Paths & Groups
    escalation_paths: Optional[str] = None
    notification_groups: Optional[str] = None
    
    # Users
    assigned_users: Optional[str] = None
    touched_by: Optional[str] = None
    closed_by: Optional[str] = None
    
    # Comments
    cs_soc_last_comment: Optional[str] = None
    customer_last_comment: Optional[str] = None
    
    # Response Actions
    response_action: Optional[str] = None
    action_target: Optional[str] = None
    target_type: Optional[str] = None
    action_provider: Optional[str] = None
    executed_date: Optional[datetime] = None
    executed_by: Optional[str] = None
    response_action_status: Optional[str] = None
    
    # Response Times (stored as timedelta for calculations)
    cs_soc_ttr: Optional[timedelta] = None
    cs_soc_ttd: Optional[timedelta] = None
    customer_ttr: Optional[timedelta] = None
    customer_ttd: Optional[timedelta] = None
    
    # MITRE ATT&CK
    mitre_tactic_id: Optional[str] = None
    mitre_tactic_name: Optional[str] = None
    mitre_technique_id: Optional[str] = None
    mitre_technique_name: Optional[str] = None
    
    # Vendor Severity
    vendor_severity: Optional[str] = None


# Column mapping: Excel column name -> Incident field name
# Includes aliases for different export formats (standard CORR, Burlington, etc.)
COLUMN_MAPPING = {
    # Identifiers
    "Incident Id": "incident_id",
    "Incident ID": "incident_id",  # Burlington format (uppercase ID)
    "Incident URL": "incident_url",
    "Vendor Incident Id": "vendor_incident_id",
    "Vendor Incident ID": "vendor_incident_id",  # Burlington format
    "Vendor Incident URL": "vendor_incident_url",
    "Incident Title": "incident_title",
    
    # Organization & Product
    "Organization": "organization",
    "Product": "product",
    "Deployment Status": "deployment_status",
    
    # Escalation Info
    "Initial Escalation Method": "initial_escalation_method",
    "Playbook URL": "playbook_url",
    
    # Status & Classification
    "Current Status": "current_status",
    "CS SOC Verdict": "cs_soc_verdict",
    "Current Priority": "current_priority",
    "Current Category": "current_category",
    
    # Timestamps - Standard CORR format
    "Created Datetime UTC": "created_datetime_utc",
    "Created Datetime (US/Central)": "created_datetime_local",
    "Last Updated Datetime UTC": "last_updated_datetime_utc",
    "Last Updated Datetime (US/Central)": "last_updated_datetime_local",
    "Escalated Datetime UTC": "escalated_datetime_utc",
    "Escalated Datetime (US/Central)": "escalated_datetime_local",
    "Closed Datetime UTC": "closed_datetime_utc",
    "Closed Datetime (US/Central)": "closed_datetime_local",
    
    # Timestamps - Burlington format (parentheses around UTC)
    "Created Datetime (UTC)": "created_datetime_utc",
    "Last Updated Datetime (UTC)": "last_updated_datetime_utc",
    "Escalated Datetime (UTC)": "escalated_datetime_utc",
    "Closed Datetime (UTC)": "closed_datetime_utc",
    
    # Timestamps - Burlington User TZ format (maps to local)
    "Created Datetime (User TZ - US/Eastern)": "created_datetime_local",
    "Last Updated Datetime (User TZ - US/Eastern)": "last_updated_datetime_local",
    "Escalated Datetime (User TZ - US/Eastern)": "escalated_datetime_local",
    "Closed Datetime (User TZ - US/Eastern)": "closed_datetime_local",
    
    # Paths & Groups
    "Escalation Paths": "escalation_paths",
    "Escalation Path": "escalation_paths",  # Burlington format (singular)
    "Notification Groups": "notification_groups",
    
    # Users
    "Assigned Users": "assigned_users",
    "Touched By": "touched_by",
    "Closed By": "closed_by",
    
    # Comments
    "CS SOC Last Comment": "cs_soc_last_comment",
    "Customer Last Comment": "customer_last_comment",
    
    # Response Actions
    "Response Action": "response_action",
    "Action Target": "action_target",
    "Target Type": "target_type",
    "Action Provider": "action_provider",
    "Executed Date": "executed_date",
    "Executed By": "executed_by",
    "Response Action Status": "response_action_status",
    
    # Response Times
    "CS SOC TTR (hh:mm)": "cs_soc_ttr",
    "CS SOC TTD (hh:mm)": "cs_soc_ttd",
    "Customer TTR (hh:mm)": "customer_ttr",
    "Customer TTD (hh:mm)": "customer_ttd",
    
    # MITRE ATT&CK
    "MITRE Tactic Id": "mitre_tactic_id",
    "MITRE Tactic ID": "mitre_tactic_id",  # Burlington format
    "MITRE Tactic Name": "mitre_tactic_name",
    "MITRE Technique Id": "mitre_technique_id",
    "MITRE Technique ID": "mitre_technique_id",  # Burlington format
    "MITRE Technique Name": "mitre_technique_name",
    
    # Vendor Severity
    "Vendor Severity": "vendor_severity",
}

# Fields that should be parsed as datetime
DATETIME_FIELDS = {
    "created_datetime_utc",
    "created_datetime_local",
    "last_updated_datetime_utc",
    "last_updated_datetime_local",
    "escalated_datetime_utc",
    "escalated_datetime_local",
    "closed_datetime_utc",
    "closed_datetime_local",
    "executed_date",
}

# Fields that should be parsed as time duration (hh:mm format)
DURATION_FIELDS = {
    "cs_soc_ttr",
    "cs_soc_ttd",
    "customer_ttr",
    "customer_ttd",
}


def parse_datetime(value) -> Optional[datetime]:
    """Parse a datetime value from Excel.
    
    Handles both datetime objects (from openpyxl) and ISO format strings.
    
    Args:
        value: The value to parse (datetime, string, or None)
        
    Returns:
        Parsed datetime or None if parsing fails
    """
    if value is None:
        return None
    
    if isinstance(value, datetime):
        return value
    
    if isinstance(value, str):
        # Try ISO format first
        try:
            # Handle timezone-aware strings
            if '+' in value or value.endswith('Z'):
                # Remove timezone info for naive datetime
                value = re.sub(r'[+-]\d{2}:\d{2}$', '', value)
                value = value.replace('Z', '')
            return datetime.fromisoformat(value)
        except ValueError:
            pass
        
        # Try common formats
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    
    return None


def parse_duration(value) -> Optional[timedelta]:
    """Parse a duration value in hh:mm format to timedelta.
    
    Args:
        value: The value to parse (string like "1:07" or "0:41")
        
    Returns:
        Parsed timedelta or None if parsing fails
    """
    if value is None:
        return None
    
    if isinstance(value, timedelta):
        return value
    
    if isinstance(value, str):
        # Handle hh:mm format
        match = re.match(r'^(\d+):(\d{2})$', value.strip())
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2))
            return timedelta(hours=hours, minutes=minutes)
        
        # Handle h:mm format (single digit hour)
        match = re.match(r'^(\d):(\d{2})$', value.strip())
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2))
            return timedelta(hours=hours, minutes=minutes)
    
    return None


def load_excel_file(file_path: Path) -> List[Incident]:
    """Load and parse a single Excel file into Incident records.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of Incident records
        
    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If required columns are missing
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    # Build column index from header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    column_index = {}
    datetime_columns_found = []
    for col_idx, header in enumerate(header_row):
        if header and header in COLUMN_MAPPING:
            field_name = COLUMN_MAPPING[header]
            column_index[col_idx] = field_name
            if field_name in DATETIME_FIELDS:
                datetime_columns_found.append(header)
    
    logger.info(f"  DateTime columns found: {datetime_columns_found}")
    
    # Validate required columns
    required_columns = {"organization", "current_status", "current_priority"}
    found_fields = set(column_index.values())
    missing = required_columns - found_fields
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    
    # Parse data rows
    incidents = []
    dates_parsed = {"created": 0, "escalated": 0, "closed": 0}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if all(cell is None for cell in row):
            continue
        
        # Build incident from row
        incident_data = {}
        for col_idx, field_name in column_index.items():
            if col_idx < len(row):
                value = row[col_idx]
                
                # Parse special field types
                if field_name in DATETIME_FIELDS:
                    value = parse_datetime(value)
                    # Track successful datetime parsing
                    if value is not None:
                        if "created" in field_name:
                            dates_parsed["created"] += 1
                        elif "escalated" in field_name:
                            dates_parsed["escalated"] += 1
                        elif "closed" in field_name:
                            dates_parsed["closed"] += 1
                elif field_name in DURATION_FIELDS:
                    value = parse_duration(value)
                
                incident_data[field_name] = value
        
        incidents.append(Incident(**incident_data))
    
    wb.close()
    
    logger.info(f"  Parsed {len(incidents)} incidents")
    logger.info(f"  Dates parsed - created: {dates_parsed['created']}, escalated: {dates_parsed['escalated']}, closed: {dates_parsed['closed']}")
    
    # Log sample of dates found
    if incidents:
        sample_dates = []
        for inc in incidents[:5]:
            dt = inc.escalated_datetime_utc or inc.created_datetime_utc or inc.closed_datetime_utc
            if dt:
                sample_dates.append(dt.strftime("%Y-%m-%d"))
        if sample_dates:
            logger.info(f"  Sample dates: {sample_dates}")
    
    return incidents


def load_multiple_periods(file_paths: List[Path]) -> Tuple[List[List[Incident]], str]:
    """Load multiple Excel files representing different reporting periods.
    
    The last file in the list is treated as the current period.
    Earlier files provide historical comparison data for trend charts.
    
    Args:
        file_paths: List of paths to Excel files (ordered chronologically)
        
    Returns:
        Tuple of:
        - List of incident lists (one per period)
        - Client name extracted from the data
        
    Raises:
        ValueError: If files are empty or have inconsistent client names
    """
    if not file_paths:
        raise ValueError("At least one Excel file is required")
    
    all_periods = []
    client_names = set()
    
    for file_path in file_paths:
        incidents = load_excel_file(file_path)
        all_periods.append(incidents)
        
        # Extract client names from this period
        for incident in incidents:
            if incident.organization:
                client_names.add(incident.organization)
    
    # Validate client name consistency
    if len(client_names) == 0:
        raise ValueError("No organization names found in the data")
    elif len(client_names) > 1:
        # Use the most common one, but warn
        import logging
        logging.warning(f"Multiple organizations found in data: {client_names}")
    
    # Get primary client name (first non-None organization from current period)
    current_period = all_periods[-1]
    client_name = None
    for incident in current_period:
        if incident.organization:
            client_name = incident.organization
            break
    
    if client_name is None and client_names:
        client_name = next(iter(client_names))
    
    return all_periods, client_name or "Unknown Client"


def get_period_date_range(incidents: List[Incident]) -> Tuple[Optional[datetime], Optional[datetime]]:
    """Extract the date range covered by a list of incidents.
    
    Args:
        incidents: List of Incident records
        
    Returns:
        Tuple of (earliest_date, latest_date) or (None, None) if no dates found
    """
    dates = []
    for incident in incidents:
        if incident.created_datetime_utc:
            dates.append(incident.created_datetime_utc)
        if incident.escalated_datetime_utc:
            dates.append(incident.escalated_datetime_utc)
    
    if not dates:
        return None, None
    
    return min(dates), max(dates)


def validate_excel_structure(file_path: Path) -> List[str]:
    """Validate that an Excel file has the expected column structure.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of warning/error messages (empty if valid)
    """
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl is required. Install with: pip install openpyxl"]
    
    file_path = Path(file_path)
    if not file_path.exists():
        return [f"File not found: {file_path}"]
    
    messages = []
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        found_columns = {h for h in header_row if h is not None}
        expected_columns = set(COLUMN_MAPPING.keys())
        
        missing = expected_columns - found_columns
        extra = found_columns - expected_columns
        
        if missing:
            messages.append(f"Missing columns: {', '.join(sorted(missing))}")
        if extra:
            messages.append(f"Unexpected columns (ignored): {', '.join(sorted(extra))}")
        
        # Check for data rows
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_count += 1
        
        if row_count == 0:
            messages.append("No data rows found in file")
        else:
            messages.append(f"Found {row_count} data rows")
        
        wb.close()
        
    except Exception as e:
        messages.append(f"Error reading file: {e}")
    
    return messages
